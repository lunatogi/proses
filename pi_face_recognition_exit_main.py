#!/usr/bin/env python3

# USAGE
# sudo python3 pi_face_recognition_exit_main.py --cascade haarcascade_frontalface_default.xml --encodings encodings.json

# import the necessary packages
from __future__ import print_function  
import msgbox
from imutils.video import VideoStream
from imutils.video import FPS
from openpyxl import load_workbook
import openpyxl
from googleapiclient.discovery import build  
from httplib2 import Http  
from oauth2client import file, client, tools  
from oauth2client.service_account import ServiceAccountCredentials
import RPi.GPIO as GPIO
from mfrc522 import SimpleMFRC522
#import spreadsheet as sheet
import tkinter as tk
import face_recognition
import argparse
import imutils
import pickle
import json
import time
import chardet
import cv2
import os
import datetime
import xlsxwriter
import xlrd
import mesaagee

root1=tk.Tk()
root1.withdraw()

root2=tk.Tk()
root2.withdraw()

root3=tk.Tk()
root3.withdraw()

MY_SPREADSHEET_ID = '1y-J0WfXE7Yc1T4_gV8P7SmNAb-qUffZR0y-cIGodcRc'

formenConf = False
run = True

endDateData = []
endDataSheet = []

def FormenTrue():
    global formenConf
    formenConf = True

def FormenFalse():
    global formenConf
    formenConf = False

def ReadCard():               #Reads data from rfid reader
    GPIO.setwarnings(False)     
    reader = SimpleMFRC522()
        
    try:
        id, text = reader.read()
        return text
    except:
        return None
    finally:
        GPIO.cleanup()
    
def UpdateSheet(date, name, surname, tc, job, comp, enter, f):
    SCOPES = 'https://www.googleapis.com/auth/spreadsheets'
    creds = ServiceAccountCredentials.from_json_keyfile_name( 
            'Proses-e8298344399a.json', SCOPES)
    service = build('sheets', 'v4', http=creds.authorize(Http()))

    # Call the Sheets API, append the next row of sensor data
    # values is the array of rows we are updating, its a single row
    fullName = name + " " + surname
    values = [ [date, fullName, tc, job, comp, enter, f] ]
    body = { 'values': values }
    # call the append API to perform the operation
    result = service.spreadsheets().values().append(
                spreadsheetId=MY_SPREADSHEET_ID, 
                range='nçıkış' + '!A1:F1',
                valueInputOption='USER_ENTERED', 
                insertDataOption='INSERT_ROWS',
                body=body).execute()  
    
def WriteOnExcel():         #İlki row ikincisi column
    global endDateData
    _date = datetime.datetime.now().date()
    loc = ("/media/pi/BB59-0061/Puantaj.xlsx")
    wb = load_workbook(filename = "/media/pi/BB59-0061/Puantaj.xlsx")
    ws = wb.get_active_sheet()
    actualRow = 0
    actualColumn = 0
    row = 1
    column = 7
    
    
    
    while(ws.cell(row = row,column = column).value != None):          #Column bulmak
        excelDate = ws.cell(row = row, column = column).value
        #excelDateCon = datetime.datetime.strptime(excelDate, '%Y.%m.%d')
        print("date :"+str(excelDate.date()))
        if(excelDate.date() == _date):
            print("Column: "+str(column))
            actualColumn = column
            break;
        else:
            column += 1
    
    for item in endDateData:
        if item["Gerçek çıkış saati"] == "":
            continue
        _name = item["İsim"]
        print(_name)
        _surname = item["Soyisim"]
        _tc = item["Tc"]
        print(_tc)
        _comp = item["Şantiye"]
        row = 2
        column = 3
        exist = False
        while(ws.cell(row = row,column = column).value != None):          #Row'u bulmak
            #print(sheet.cell(row = row,column = column).value)
            #print("Tc enc :"+_tc)
            #print("Tc puan :"+str(sheet.cell(row = row, column = column).value))
            if(str(ws.cell(row = row, column = column).value) == _tc):
                print("Row: "+str(row))
                exist = True
                actualRow = row
                break;
            else:
                row += 2
        
        #wb = load_workbook(filename = "/home/pi/Desktop/Belgeler/Puantaj1.xlsx")
        #ws = wb.get_active_sheet()
        fullName = _name+" "+_surname
        if exist == False:
            actualRow = row
            ws.cell(row=actualRow, column=2).value = fullName
            ws.cell(row=actualRow+1, column=2).value = fullName
            ws.cell(row=actualRow, column=3).value = _tc
            ws.cell(row=actualRow+1, column=3).value = _tc
            ws.cell(row=actualRow, column=5).value = _comp
            ws.cell(row=actualRow+1, column=5).value = _comp
            ws.cell(row=actualRow, column=6).value = "MESAİ"
            ws.cell(row=actualRow+1, column=6).value = "GÜN"
        print("actualrow :"+str(actualRow))
        try:
            valEnt = datetime.datetime.strptime(item["Giriş saati"], '%H:%M')
            valOut = datetime.datetime.strptime(item["Çıkış saati"], '%H:%M')
            sumHour =  valOut - valEnt
            _sumHour = sumHour
        except:
            try:
                valEnt = datetime.datetime.strptime(item["Gerçek giriş saati"], '%H:%M')
                valOut = datetime.datetime.strptime(item["Gerçek çıkış saati"], '%H:%M')
                sumHour =  valOut - valEnt
                _sumHour = sumHour
            except:
                _sumHour = "-"
        

                

        print(str(actualRow))
        print(str(actualColumn))
        ws.cell(row=actualRow,column=actualColumn).value = "0"
        try:
            if(_sumHour > datetime.timedelta(hours = 4)):
                ws.cell(row=actualRow+1,column=actualColumn).value = "1"
            else:
                ws.cell(row=actualRow+1,column=actualColumn).value = "0.5"
        except:
            ws.cell(row=actualRow+1, column=actualColumn).value = "-"
    
    wb.save("/media/pi/BB59-0061/Puantaj.xlsx")
    

def EndProcess():
    global run
    print("girdiiii")
    global endDateData
    global endDataSheet
    enter_currDate = str(datetime.datetime.now().date())
    with open(enter_currDate+".json", "w") as d:
        json.dump(endDateData, d, ensure_ascii = False)

    # workbook = xlsxwriter.Workbook('/../../../../media/pi/LUNAWINDOWS/'+enter_currDate+'.xlsx')         # Bu çalışırken programın nerede çalıştığı önemli
    workbook = xlsxwriter.Workbook(''+enter_currDate+'.xlsx')
    worksheet = workbook.add_worksheet()

    #sheet.UpdateJson(item)

    row = 0
    col = 0
    worksheet.write(row, col, "Isim")
    worksheet.write(row, col+1, "Soyisim")
    worksheet.write(row, col+2, "Tc Numarası")
    worksheet.write(row, col+3, "Meslek")
    worksheet.write(row, col+4, "Şantiye")
    worksheet.write(row, col+5, "Giris Saati")
    worksheet.write(row, col+6, "Gerçek Giris Saati")
    worksheet.write(row, col+7, "Cikis Saati")
    worksheet.write(row, col+8, "Gerçek Cikis Saati")
    worksheet.write(row, col+9, "Toplam Saat")
    worksheet.write(row, col+10, "Formen Onayı")
    row = 1
    col = 0
    for item in endDateData:
        #if item["Gerçek çıkış saati"] == "":
            #continue
        print("value: "+item["İsim"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["İsim"])
        #row = 1
        col = 1
        print("value: "+item["Soyisim"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Soyisim"])
        #row = 1
        col = 2 
        print("value: "+item["Tc"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Tc"])
        #row = 1
        col = 3 
        print("value: "+item["Meslek"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Meslek"])
        #row = 1
        col = 4 
        print("value: "+item["Şantiye"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Şantiye"])
        #row = 1
        col = 5            
        print("value: "+item["Giriş saati"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Giriş saati"])
        #row = 1
        col = 6            
        print("value: "+item["Gerçek giriş saati"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Gerçek giriş saati"])
        #row = 1
        col = 7
        print("value: "+item["Çıkış saati"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Çıkış saati"])
        #row = 1
        col = 8
        print("value: "+item["Gerçek çıkış saati"]+" r,c: "+str(row)+","+str(col))
        worksheet.write(row, col, item["Gerçek çıkış saati"])
        #row = 1
        col = 9    
        sumHour = 0              
        try:
            valEnt = datetime.datetime.strptime(item["Giriş saati"], '%H:%M')
            valOut = datetime.datetime.strptime(item["Çıkış saati"], '%H:%M')
            sumHour =  valOut - valEnt
            print(str(sumHour))
            worksheet.write(row, col, str(sumHour))
        except:
            try:
                valEnt = datetime.datetime.strptime(item["Gerçek giriş saati"], '%H:%M')
                valOut = datetime.datetime.strptime(item["Gerçek çıkış saati"], '%H:%M')
                sumHour =  valOut - valEnt
                print(str(sumHour))
                worksheet.write(row, col, str(sumHour))
            except:
                worksheet.write(row, col, "-")
        col = 10
        worksheet.write(row, col, item["Formen"])
        row += 1
        col = 0
        name = item["İsim"]
        surname = item["Soyisim"]
        tc = item["Tc"]
        comp = item["Şantiye"]
        date = datetime.datetime.now().date()
        #dateStr = str(day)
        sumInt = int(str(sumHour)[0])
        
    workbook.close()
    WriteOnExcel()

    try:
        sheetJson = []
        try:
            with open("nesheet.json", "r") as read_file:
                sheetJson = json.load(read_file)
            for item in sheetJson:
                endDataSheet.append(item)
        except:
            pass
        for item in endDataSheet:
            pass
            #UpdateSheet(enter_currDate, item["İsim"], item["Soyisim"], item["Tc"], item["Meslek"], item["Şantiye"], item["Gerçek çıkış saati"], item["Formen"])
        try:
            os.remove("nesheet.json")
        except:
            pass
    except:
        #tk.messagebox.showinfo("Hata", "Veriler internet problemi nedeniyle Google Sheets'e atılamadı.")
        #root2.lift()
        #with open("nesheet.json", "w") as d:
            #json.dump(endDataSheet, d, ensure_ascii = False)
        pass
    try:
        oldDate = str(datetime.datetime.now().date() - datetime.timedelta(days = 1))
        print(oldDate)
        os.remove(oldDate+".json")
    except:
        pass
    run = False

def FaceRecOut():
    global formenConf
    global run
    global endDateData
    global endDataSheet
    formenOnay = False
    formenOnay = formenConf
    # construct the argument parser and parse the arguments
    ap = argparse.ArgumentParser()
    ap.add_argument("-c", "--cascade", required=False,
        help = "path to where the face cascade resides")
    ap.add_argument("-e", "--encodings", required=False,
        help="path to serialized db of facial encodings")
    ap.add_argument("-i", "--dataset", required=False,
        help="path to input directory of faces + images")
    ap.add_argument("-d", "--detection-method", type=str, default="hog",
        help="face detection model to use: either `hog` or `cnn`")
    args = vars(ap.parse_args())

    # load the known faces and embeddings along with OpenCV's Haar
    # cascade for face detection
    print("[INFO] loading encodings + face detector...")
    #data = json.loads(open(args["encodings"]).read())
    detector = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
    
    currDate = str(datetime.datetime.now().date())
    
    run = True
    try:
        with open(currDate+".json", "r") as read_file:
            endDateData = json.load(read_file)
        print(endDateData)
    except:
        msgbox.messagebox("Çıkış yapmak için en az bir kişinin giriş yapması gerekmektedir.")
        #tk.messagebox.showinfo("Uyarı", "Çıkış yapmak için en az bir kişinin giriş yapması gerekmektedir.")
        #root3.lift()
        run = False
        return;
    if formenOnay == True:
        print("deneme")
    
    # initialize the video stream and allow the camera sensor to warm up
    print("[INFO] starting video stream...")
    # vs = VideoStream(src=0).start()
    vs = VideoStream(usePiCamera=True).start()
    time.sleep(2.0)

    # start the FPS counter
    fps = FPS().start()
    
    writeDataBool = True
    rfidNumber = 00
    
    
    # loop over frames from the video file stream
    enter_currTime = ''
    
    #result = tk.messagebox.askquestion("Uyarı", "Başlamak için 'yes'e çıkmak için 'no'ya basınız.")
    #root1.lift()
   
    result=mesaagee.message("Başlamak için bir kart okutun, çıkmak için 'Bitir'e basın.",00)
    if result == 'yes':
        run = False
    else:
        run = True
    while run == True:
        firstExitTime = datetime.datetime.strptime('16:30', '%H:%M').time()
        lastExitTime = datetime.datetime.strptime('17:30', '%H:%M').time()
        currDateTimeForExit = datetime.datetime.now().time()
        enter_currTime = '17:00'
        if currDateTimeForExit >= firstExitTime:
            if currDateTimeForExit <= lastExitTime:
                enter_currTime = '17:00'
            else:
                enter_currTime = "-"
                #print("Giriş saatini geçtiniz lütfen formen onayını alın.")
                #run = False
                #return;
        else:
            enter_currTime = "-"
            #print("Giriş saatini geçtiniz lütfen formen onayını alın.")
            #run = False
            #return;
        # print(str(enter_currTime))
        # grab the frame from the threaded video stream and resize it
        # to 500px (to speedup processing)
        frame = vs.read()
        if writeDataBool:
        
            frame = imutils.resize(frame, width=500)
    
        # convert the input frame from (1) BGR to grayscale (for face
        # detection) and (2) from BGR to RGB (for face recognition)
            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        # detect faces in the grayscale frame
            rects = detector.detectMultiScale(gray, scaleFactor=1.1, 
                minNeighbors=5, minSize=(150, 150),
                flags=cv2.CASCADE_SCALE_IMAGE)

        # OpenCV returns bounding box coordinates in (x, y, w, h) order
        # but we need them in (top, right, bottom, left) order, so we
        # need to do a bit of reordering
            boxes = [(y, x + w, y + h, x) for (x, y, w, h) in rects]

        # compute the facial embeddings for each face bounding box
            encodings = face_recognition.face_encodings(rgb, boxes)
            names = []
        
        name = "Unknown"
        surname = "Surname"
        tc = "Tc"
        job = "Job"
        comp = "Comp"    
            
        readerValue = str(ReadCard())
        if readerValue != "None":
            rfidNumber = ReadCard()
        
        
        print("rfid: "+str(rfidNumber))
        actualIndex = None
        
        known_faces = []
        database = []
        with open("database.json", "r") as read_file:
            database = json.load(read_file)
        counter = 0
        try:
            if rfidNumber != 00:
                for i in database["tc"]:
                    print("i in database: "+str(i))
                    if int(i) == int(rfidNumber):
                        print("Buldu")
                        actualIndex = counter
                    else:
                        counter += 1
        except:
            pass
        # loop over the facial embeddings
        for encoding in encodings:
            databasePerson = []
            known_faces = []
            #try:
            print("actl Index: "+str(actualIndex))
            try:
                databasePerson.append(database["encodings"][actualIndex])
                databasePerson.append(database["encodings"][actualIndex+1])
                databasePerson.append(database["encodings"][actualIndex+2])
                databasePerson.append(database["encodings"][actualIndex+3])
            except:
                pass
            known_faces = databasePerson
            print("faces"+str(known_faces))
            matches = []
            if actualIndex != None:
                matches = face_recognition.compare_faces(known_faces,
                    encoding)
            else:
                matches = [False]
                
                
            trueCounter = 0
            for vari in matches:
                if vari == True:
                    trueCounter += 1
            print("trueCounter: "+str(trueCounter))
            # check to see if we have found a match
            if trueCounter >= 3:
            
                # find the indexes of all matched faces then initialize a
                # dictionary to count the total number of times each face
                # was matched

                # loop over the matched indexes and maintain a count for
                # each recognized face face 
                #encodeCounter -= 1

                name = database["names"][actualIndex]
                surname = database["surname"][actualIndex]
                job = database["job"][actualIndex]
                comp = database["comp"][actualIndex]
                tc = database["tc"][actualIndex]
                    
                print("name: "+name)
                print("comp: "+comp)

                # determine the recognized face with the largest number
                # of votes (note: in the event of an unlikely tie Python
                # will select first entry in the dictionary)
                actualExitTime = str(datetime.datetime.now().time())[0:5]
                if writeDataBool:
                    
                    # enter_currTime = str(datetime.datetime.now().time())
                    # update the list of names
                    names.append(name)
                    exist = False
                    alreadyOut = False
                    for item in endDateData:
                        print(enter_currTime)
                        if item["Tc"] == tc:
                            exist = True
                            if item["Çıkış saati"] == "":
                                item["Çıkış saati"] = enter_currTime
                                item["Gerçek çıkış saati"] = actualExitTime
                                expDateData = {"Çıkış saati":enter_currTime, "Gerçek çıkış saati":actualExitTime, "Giriş saati": "", "Gerçek giriş saati": "",  "İsim": name, "Soyisim": surname, "Tc": tc, "Meslek": job, "Şantiye": comp, "Formen": ""}
                                endDataSheet.append(expDateData)
                                
                            else: 
                                alreadyOut = True
                            break
                    if exist == True:
                        if alreadyOut == True:
                            #result1 = tk.messagebox.askquestion("Uyarı", "Zaten çıkış yapmış bulunmaktasınız.")
                            #root1.lift()
                            
                            result=mesaagee.message("Zaten giriş yapmış bulunmaktasınız.\n Devam etmek için yeni bir kart okutun, bitirmek için 'Bitir'e basınız.",rfidNumber)
                            if result == 'yes':
                                EndProcess()
                            else:
                                pass   
                        else:
                            #result = tk.messagebox.askquestion("Bilgi", "İsim: "+name+" "+surname+"- Gerçek Çıkış Saati: "+actualExitTime+" - Devam edilsin mi?")
                            #root1.lift
                         
                            result=mesaagee.message("Bilgi, İsim: "+name+" "+surname+"- Gerçek Çıkış Saati: "+actualExitTime+" -\n Devam etmek için yeni kart okutun, bitirmek için 'Bitir'e basın.",rfidNumber)
                            if result == 'yes':
                                EndProcess()
                            else:
                                pass
                    else:
                        #tk.messagebox.showinfo("Uyarı", "Çıkış yapmak için önce giriş yapmanız gerekmektedir.")
                        #root1.lift()
                       
                        result=mesaagee.message("Bilgi, Çıkış yapmak için önce giriş yapmanız gerekmektedir.\n Devam etmek için yeni bir kart okutun, bitirmek için 'Bitir'e basın.",rfidNumber)
                        if result == 'yes':
                            EndProcess()
                        else:
                            pass
                    # loop over the recognized faces
                    for ((top, right, bottom, left), name) in zip(boxes, names):
                        # draw the predicted face name on the image
                        cv2.rectangle(frame, (left, top), (right, bottom),
                            (0, 255, 0), 2)
                        y = top - 15 if top - 15 > 15 else top + 15
                        cv2.putText(frame, name, (left, y), cv2.FONT_HERSHEY_SIMPLEX,
                            0.75, (0, 255, 0), 2)

        # display the image to our screen
        cv2.imshow("Frame", frame)
        key = cv2.waitKey(1) & 0xFF


        if key == ord("k"):
            writeDataBool = True
    
        # if the `q` key was pressed, break from the loop
        if key == ord("q"):
            enter_currDate = str(datetime.datetime.now().date())
            with open(enter_currDate+".json", "w") as d:
                json.dump(endDateData, d, ensure_ascii = False)
            
            # workbook = xlsxwriter.Workbook('/../../../../media/pi/LUNAWINDOWS/'+enter_currDate+'.xlsx')         # Bu çalışırken programın nerede çalıştığı önemli
            workbook = xlsxwriter.Workbook(''+enter_currDate+'.xlsx')
            worksheet = workbook.add_worksheet()
        
            row = 0
            col = 0
            worksheet.write(row, col, "Isim")
            worksheet.write(row, col+1, "Soyisim")
            worksheet.write(row, col+2, "Tc Numarası")
            worksheet.write(row, col+3, "Meslek")
            worksheet.write(row, col+4, "Şantiye")
            worksheet.write(row, col+5, "Giris Saati")
            worksheet.write(row, col+6, "Gerçek Giris Saati")
            worksheet.write(row, col+7, "Cikis Saati")
            worksheet.write(row, col+8, "Gerçek Cikis Saati")
            worksheet.write(row, col+9, "Toplam Saat")
            row = 1
            col = 0
            for item in endDateData:
                worksheet.write(row, col, item["İsim"])
                #row = 1
                col = 1
                
                worksheet.write(row, col, item["Soyisim"])
                #row = 1
                col = 2 
                
                worksheet.write(row, col, item["Tc"])
                #row = 1
                col = 3 
                
                worksheet.write(row, col, item["Meslek"])
                #row = 1
                col = 4 
                
                worksheet.write(row, col, item["Şantiye"])
                #row = 1
                col = 5 
            
                worksheet.write(row, col, item["Giriş saati"])
                #row = 1
                col = 6            
            
                worksheet.write(row, col, item["Gerçek giriş saati"])
                #row = 1
                col = 7
            
                worksheet.write(row, col, item["Çıkış saati"])
                #row = 1
                col = 8
            
                worksheet.write(row, col, item["Gerçek çıkış saati"])
                #row = 1
                col = 9    
                sumHour = 0              
                try:
                    valEnt = datetime.datetime.strptime(item["Giriş saati"], '%H:%M')
                    valOut = datetime.datetime.strptime(item["Çıkış saati"], '%H:%M')
                    sumHour =  valOut - valEnt
                    print(str(sumHour))
                    worksheet.write(row, col, str(sumHour))
                except:
                    try:
                        valEnt = datetime.datetime.strptime(item["Gerçek giriş saati"], '%H:%M')
                        valOut = datetime.datetime.strptime(item["Gerçek çıkış saati"], '%H:%M')
                        sumHour =  valOut - valEnt
                        print(str(sumHour))
                        worksheet.write(row, col, str(sumHour))       
                    except:
                        worksheet.write(row, col, str(sumHour))  
                row += 1
                col = 0
                name = item["İsim"]
                surname = item["Soyisim"]
                tc = item["Tc"]
                comp = item["Şantiye"]
                day = datetime.datetime.now().date().day
                dayStr = str(day)
                sumInt = int(str(sumHour)[0])
                print(name+" "+surname+" ")
                WriteOnExcel(name, surname, tc, comp, sumInt, dayStr)
            
            workbook.close()
            
            os.remove(enter_currDate+".json")
        
            break

        # update the FPS counter
        fps.update()

    # stop the timer and display FPS information
    fps.stop()
    print("[INFO] elasped time: {:.2f}".format(fps.elapsed()))
    print("[INFO] approx. FPS: {:.2f}".format(fps.fps()))

    # do a bit of cleanup
    cv2.destroyAllWindows()
    vs.stop()
